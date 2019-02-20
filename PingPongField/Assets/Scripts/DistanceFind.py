import cv2
import numpy as np
import os
import shutil
import math
import sys
from PIL import Image, ImageDraw
import openpyxl as excel
import re

Xdata_list_Camera1 = [] #見つけた領域を保存しておくリスト
Xdata_list_Camera2 = [] #見つけた領域を保存しておくリスト
r4 = [] #見つけた領域を保存しておくリスト
r3 = [] #見つけた領域を保存しておくリスト


CalibrationData_Number =29 #キャリブレーションに必要な写真の数

def Camera1():
    #ファイルの数を数える
    files = os.listdir('../ExperimentData')
    count = 0#ディレクトリのファイルを数えるための変数

    for file in files:
        index = re.search('.png',file) #pngのファイルを探す
        if index: #pngのファイルであれば
            count = count + 1
    DataCount = count/4 -CalibrationData_Number; #何個の追加写真があるのか数える

    #print(int(DataCount),"バウンド")

    for i in range(29 + int(DataCount)):
        #print(i+1,"枚目")
        img1 =cv2.imread("../ExperimentData/Point"+str(i+1)+"_Camera1.png")
        img2 =cv2.imread("../ExperimentData/Point29_Camera1.png")#背景は29番の画像

        img_gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
        img_gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

        th=1#閾値の決定
        mask = cv2.absdiff(img_gray1,img_gray2)

        # 差分画像を二値化してマスク画像を算出
        mask[mask < th] = 0
        mask[mask >= th] = 255

        #画像のサイズの取得
        H,W,_ = img1.shape

        #画像の表示
        #cv2.imshow('image',mask)
        #cv2.waitKey(100000)
        #cv2.destroyAllWindows()
        cv2.imwrite("../ProcessingExperimentData/Data"+str(i+1)+"_Camera1.png",mask)

        #領域を見つけ出して座標を取得する
        _,contours, _ = cv2.findContours(mask, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)

        for i in range(int(len(contours))):
            area = cv2.contourArea(contours[i])#面積を計算するよ
            #print("面積",area)
            if area > 50.0: #面積が大きくないとノイズとみなす
                rect_x ,rect_y ,rect_w, rect_h = cv2.boundingRect(contours[i])
                x_data = int(rect_x+(rect_w/2))
                y_data = int(rect_y+(rect_h/2))
                #配列にXのデータの格納
                Xdata_list_Camera1.append(((W/2)-x_data))
                #print("配列のながさは",len(Xdata_list_Camera1))
                #print("##########################")

def Camera2():

    #ファイルの数を数える
    files = os.listdir('../ExperimentData')
    count = 0#ディレクトリのファイルを数えるための変数

    for file in files:
        index = re.search('.png',file) #pngのファイルを探す
        if index: #pngのファイルであれば
            count = count + 1
    DataCount = count/4 - CalibrationData_Number; #何個の追加写真があるのか数える

    for i in range(29 + int(DataCount)):
        #print(i+1,"枚目")
        img1 =cv2.imread("../ExperimentData/Point"+str(i+1)+"_Camera2.png")
        img2 =cv2.imread("../ExperimentData/Point29_Camera2.png")#背景は29番の画像

        img_gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
        img_gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)

        th=1
        mask = cv2.absdiff(img_gray1,img_gray2)

        # 差分画像を二値化してマスク画像を算出
        mask[mask < th] = 0
        mask[mask >= th] = 255

        #画像のサイズの取得
        H,W,_ = img1.shape

        #画像の表示
        #cv2.imshow('image',mask)
        #cv2.waitKey(100000)
        #cv2.destroyAllWindows()
        cv2.imwrite("../ProcessingExperimentData/Data"+str(i+1)+"_Camera2.png",mask)

        #領域を見つけ出して座標を取得する
        _,contours, _ = cv2.findContours(mask, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)

        for j in range(int(len(contours))):
            area = cv2.contourArea(contours[j])#面積を計算するよ

            #print("面積",area)
            if area > 50.0: #面積が大きくないとノイズとみなす
                rect_x ,rect_y ,rect_w, rect_h = cv2.boundingRect(contours[j])
                x_data = int(rect_x+(rect_w/2))
                y_data = int(rect_y+(rect_h/2))
                #配列にXのデータの格納
                Xdata_list_Camera2.append(((W/2)-x_data))
                #print("配列のながさは",len(Xdata_list_Camera2))
                #print("##########################")



########################ここからプログラムのスタート

Camera1()
Camera2()

print("カメラ1のバウンド数",len(Xdata_list_Camera1))
print("カメラ2のバウンド数",len(Xdata_list_Camera2))

for i in range(len(Xdata_list_Camera1)):
    r3.append(Xdata_list_Camera1[i] * math.cos(math.radians(-45)) - Xdata_list_Camera2[i] * math.sin(math.radians(-45)))
    r4.append(Xdata_list_Camera1[i] * math.sin(math.radians(-45)) + Xdata_list_Camera2[i] * math.cos(math.radians(-45)))


###ここからエクセルのファイルの作成
if(len(Xdata_list_Camera1) == len(Xdata_list_Camera2)):#ボールの数が同じかどうかの確認
    wb = excel.Workbook()
    ws = wb.active

    for i in range(len(Xdata_list_Camera1)):#配列データをエクセルに書き込むよ
        ws["B"+str(i+2)] = r3[i]
        ws["C"+str(i+2)] = -r4[i]

for i in range(100):
    ws["A"+str(i+30)] = str(1+i)+"バウンド"


wb.save("../PingPong_Expriment1.xlsx")
